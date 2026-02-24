from dotenv import load_dotenv
load_dotenv()

import streamlit as st
import requests
import pandas as pd
import plotly.express as px
import os
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Pesquise Mais", layout="wide")

LOGO_B64 = "/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCABMANIDASIAAhEBAxEB/8QAHAABAAMBAQEBAQAAAAAAAAAAAAYHCAUEAwIB/8QARRAAAQMEAQIEBAIECQsFAAAAAQIDBAAFBhEHEiEIEzFBFCJRYTKhFSNxkQkzNTdCcoGz8BYXGCQlQ2KDsbLBUnSCkpP/xAAbAQEBAAIDAQAAAAAAAAAAAAAABQQGAQIHA//EADARAAIBAwMBBQYHAQAAAAAAAAABAgMFEQQhMRIGQWFxwTJRgZGx0RMUIjM1QqFy/9oADAMBAAIRAxEAPwDGVeq2W6fdJIjW6G/Le1voaQVED6nXoPvX1x22m73uJbRIbj/EOdJdcPZI9T+0/Qe50K0Xjtvs9jt6INt8hlpOio9YKnFf+pR9z/gaHarlnssrhmcpdMF82/cjA1utWmWyy2Uo3xpmq0hQs4G/rLZB/wC+vzJ42zVhouKsbjgHs0824r+xKVEmtCtyI/b/AFhn/wDQV6460ObLa0r169KgdVbq9m9JH2ZS+a+xr1XtBqof1Xyf3Mjy40iJIXHlsOx3kHSm3UFKk/tB7ivlWrspxiz5RbzEusYKUE6akJGnWT9Uq+n2PY1mrM8dnYvf37TO0oo0pp0DSXWz+FY/x2II9q1nX26ekec5j7yta71S17cMdM13fY41KVPeK+OlZ1FuLybumB8EtpGjH8zr6ws7/ENa6PzqfGLk8IoazWUdFRdevLEVjL3fLx3eJAqVPuVeOF4JEtshV4TcPjluo6RH8vo6Ag7/ABHe+v8AKoDRpp4Y0Wtoa6iq9CWYvOHuuHh84fIpWg/D/wCGO7ckYqMwvd/ax2wuFfwy/J8119KCQteipKUIBBHUSSek9taJkfJHhSstlwS65fjXJ0C4RLXGXIfTIZT0KCQToONqV8x7AJ6e5I71wZRlmlKUApWoPBZw5gnJ+O5FMy6BKkvwJbLbBalLaASpCiQQk9+4rMchIRIcQn0SsgfvoD50pQdzoUApU65X4qy3jJuyqypmIyu8MKfYaZe61oCenqSsaHSodYHv796nnEWG8I3fg/Ir5meWqgZXH874aN8WltaOlALXltEbe6ldjrf0+XWyBRNKUoBSlKAUpSgFKUoBX3gxJU6U3EhR3ZD7h0httJUo+/oPtX9t0KVcZzUKEyp+Q8rpQhPqT/4H39qvvAcPiYxbVrV0P3J5sh98DsBr8CPon8z6n2Ar2m0VbhN42guX6LxMPWayOmhnl9yM+V94UuVBkokw5L0Z9B2lxpZSpJ+xFfA0qSm08oy2k1hmkOHMweymzOs3ApNyhFKXVga81BB6V6+vYg/2H3rmeI20tycVhXhKR58OSGiQO5bcB9fsFJH/ANj9a5XhttchCbreXElLDgTGaJ/pkHqUR+z5f31J+f5DbHG7raiOqRLZbQPuNqJ/cn862WU5Vbf1VecfTg88qQjpr5GNDjqX+8r/AFlRYrxtk2TWdu62tuIqMtakAuPhKtpOj2q6uCcOvWHwrw1eUMJVLdZU15ToX2QF73r0/EKjHDee4nj+Dx7bd7sI0pL7qlN/Dur0Ce3dKSPzq1MSyixZQ1IdsU/4xEZSUvHylo6SrevxAb/CfT6VJp06SSknvgwe1Nxuc4VtPUp4o556Xwntvxvt5kY50wu95nb7MxZER1LiOvKd850I7LCANb9fwmqUy/i/KsVsa7xdm4aYqHEtktSAtW1enatN5LlNgxZqM9fiumAy26Mu6rSFCzgb+stkH/vr8yeNs1YaLirG44B7NNNuK/sSlRJrQrcgP2/1hn/9BXrjrQ5strSvXr0qB1Vur2b0kfZlL5r7GvVe0Gqh/VfJ/cyPLjSIkhceWw7HeQdKbdQUqT+0HuK+Va+ynGLPlFvMS6xgpQTpqQkadZP1Sr6fY9jWaszx2di9/ftM7SijSmXQNJdbP4Vj/HYgj2rWdfbp6R5zmPvK1rvVLXtwx0zXd9jjUpU94r46VnUW4vJu6YHwS2kaMfzOvrCzv8Q1ro/Op8YuTwihrNZR0VF168sRWMvd8vHd4kCpU+5V44XgkS2yFXhNw+OW6jpEfy+joCDv8Q3rr/ACqA0aaeCNFraGuoqvQlmLzh7rh4fOHyKVoPw/8AhjuxJGKjML3f2sdsLhX8Mvyfddfkgt6K0S0lIBBHUSSnoJJBI6SiVDUsABJ2k7+prxvTSt12W4rgtpLqio/+8Hjv/fRSlRz5Xl3E6f0rHuvwJMpLIUBpCEAbPf8ACKqkbHYUqWTBfO66LN2ycpwfXsKVKlAKUpQClKUApSlAK+sGJKnSm4kKO7IfcOkNtJKlH9oPcV9be1LuM5qFCZU/IeV0oQn1J/8AD7+1X3gOHxMYtq1q6H7k82Q++B2A1+BH0T+Z9T7AV7TaKtwm8bQXL9F4mHrNZHTQzy+5GfK+8KXKgyUSYcl6M+g7S40spUk/YivgaVJTaeci7SawyP8ANm4OZbZ3WbgUm5QilLqwNeagg9K9fXsQf7D71zPEbaW5OK5rwlI8+HJDRIHctuA+v2Ckj/7H66k/P8htjjd1tRHVIlstKGW9jaFE/uT99bLKcqtv6qvOPpweeVIR018jGhx1L/eV/rKixXjbJsmtzdxt3c2mO20tXkXBkuqXrtpJI0R+e/pUUpXEpdTycWe2RtmkjpYy6ks7+bbnlvFYKlT7lXjheCRLbIVeE3D45bqOkR/L6OgIO/xDeuv8qgNGmnhjRa2hrqKr0JZi84e64eHzh8ilaD8P/hju3JGKjML3f2sdsLhX8Mvyfddfkgt6K0S0lIBBHUSSnoJJBI6SiVDUsABJ2k7+prxvTSt12W4rgtpLqio/+8Hjv/fRSlRz5Xl3E6f0rHuvwJMpLIUBpCEAbPf8IqoB3oBSlKAUpSgFKUoBSlKAUpSgFKUoBSlKAUpSgFKUoBSlKAUpSgFKUoBSlKAUpSgFKUoBSlKA/9k="

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Open+Sans:wght@300;400;500;600;700;800&display=swap');

html, body, [class*="css"], .stApp, p, span, div, td, th, a {
    font-family: 'Open Sans', sans-serif !important;
}
.stApp {
    background: #0a0a0f;
    background-image:
        radial-gradient(ellipse 80% 50% at 50% -10%, rgba(99,102,241,0.18) 0%, transparent 60%),
        radial-gradient(ellipse 40% 30% at 80% 80%, rgba(236,72,153,0.10) 0%, transparent 50%);
}
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding: 2rem 3rem 4rem; max-width: 1400px; }

/* ── Navbar ── */
.navbar {
    display: flex; align-items: center; justify-content: space-between;
    padding: 1rem 0 0.5rem; margin-bottom: 0.5rem;
    border-bottom: 1px solid rgba(255,255,255,0.06);
}
.navbar-logo img { height: 42px; width: auto; }
.navbar-tag {
    font-size: 0.65rem; font-weight: 700; letter-spacing: 0.2em; text-transform: uppercase;
    color: #818cf8; background: rgba(99,102,241,0.1); border: 1px solid rgba(99,102,241,0.22);
    padding: 0.28rem 0.9rem; border-radius: 100px;
}

/* ── Hero ── */
.hero-header { text-align: center; padding: 2.5rem 0 1.8rem; }
.hero-title {
    font-size: clamp(2rem, 5vw, 3.4rem); font-weight: 800;
    line-height: 1.1; letter-spacing: -0.02em; color: #f8fafc; margin: 0 0 0.65rem;
}
.hero-title span {
    background: linear-gradient(135deg, #818cf8 0%, #ec4899 100%);
    -webkit-background-clip: text; -webkit-text-fill-color: transparent; background-clip: text;
}
.hero-sub { font-size: 0.98rem; color: #94a3b8; font-weight: 300; margin: 0; }

/* ── Search bar: match height precisely ── */
.search-wrap {
    background: rgba(255,255,255,0.03); border: 1px solid rgba(255,255,255,0.08);
    border-radius: 16px; padding: 1.4rem 1.8rem; margin: 1.2rem 0 0.4rem;
}

/* Label alignment helper — invisible label same size as real label */
.btn-label-spacer {
    font-size: 0.72rem; line-height: 1.2; color: transparent;
    margin-bottom: 6px; display: block; pointer-events: none; user-select: none;
}

.stTextInput > div > div > input {
    background: rgba(255,255,255,0.05) !important;
    border: 1px solid rgba(255,255,255,0.12) !important;
    border-radius: 10px !important; color: #f1f5f9 !important;
    font-size: 0.95rem !important; padding: 0 1rem !important;
    height: 46px !important; min-height: 46px !important; max-height: 46px !important;
    transition: all 0.2s ease !important; box-sizing: border-box !important;
}
.stTextInput > div > div > input:focus {
    border-color: rgba(99,102,241,0.6) !important;
    box-shadow: 0 0 0 3px rgba(99,102,241,0.15) !important;
    background: rgba(99,102,241,0.07) !important;
}
.stTextInput > div > div > input::placeholder { color: #475569 !important; }
.stTextInput label {
    color: #94a3b8 !important; font-size: 0.72rem !important; font-weight: 600 !important;
    letter-spacing: 0.1em !important; text-transform: uppercase !important;
    margin-bottom: 6px !important; display: block !important;
}

/* Button exactly 46px tall, full column width */
div[data-testid="stButton"] > button {
    background: linear-gradient(135deg, #6366f1 0%, #8b5cf6 100%) !important;
    color: white !important; font-weight: 700 !important; font-size: 0.88rem !important;
    letter-spacing: 0.04em !important; border: none !important; border-radius: 10px !important;
    width: 100% !important; height: 46px !important; min-height: 46px !important;
    max-height: 46px !important; padding: 0 1rem !important;
    box-sizing: border-box !important;
    transition: all 0.25s cubic-bezier(0.4,0,0.2,1) !important;
    box-shadow: 0 4px 18px rgba(99,102,241,0.35) !important;
    display: flex !important; align-items: center !important; justify-content: center !important;
    margin-top: 0 !important;
}
div[data-testid="stButton"] > button:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 26px rgba(99,102,241,0.5) !important;
    background: linear-gradient(135deg, #818cf8 0%, #a78bfa 100%) !important;
}
div[data-testid="stButton"] > button:active { transform: translateY(0) !important; }

/* Export button — green tones */
.export-btn div[data-testid="stButton"] > button {
    background: linear-gradient(135deg, #059669 0%, #10b981 100%) !important;
    box-shadow: 0 4px 18px rgba(16,185,129,0.3) !important;
}
.export-btn div[data-testid="stButton"] > button:hover {
    background: linear-gradient(135deg, #10b981 0%, #34d399 100%) !important;
    box-shadow: 0 8px 26px rgba(16,185,129,0.45) !important;
}

.stCaption { color: #3f4858 !important; font-size: 0.72rem !important; }

/* ── Section heading ── */
.section-heading {
    font-size: 1.1rem; font-weight: 700; color: #f1f5f9;
    letter-spacing: -0.01em; margin: 2rem 0 0.8rem;
    display: flex; align-items: center; gap: 0.55rem;
}
.section-heading::after {
    content: ''; flex: 1; height: 1px;
    background: linear-gradient(90deg, rgba(99,102,241,0.4), transparent);
}
.results-badge {
    display: inline-flex; align-items: center; gap: 0.4rem;
    font-size: 0.72rem; font-weight: 600; color: #818cf8;
    background: rgba(99,102,241,0.1); border: 1px solid rgba(99,102,241,0.2);
    padding: 0.24rem 0.72rem; border-radius: 100px;
    letter-spacing: 0.03em; margin-bottom: 0.9rem;
}

/* ── Table ── */
.product-table-wrapper {
    background: rgba(255,255,255,0.02); border: 1px solid rgba(255,255,255,0.07);
    border-radius: 14px; overflow: hidden; margin-bottom: 1rem;
}
.product-table-wrapper table { width: 100%; border-collapse: collapse; }
.product-table-wrapper thead tr {
    background: rgba(99,102,241,0.1); border-bottom: 1px solid rgba(99,102,241,0.2);
}
.product-table-wrapper thead th {
    padding: 0.8rem 1rem; text-align: left; font-size: 0.63rem; font-weight: 700;
    letter-spacing: 0.12em; text-transform: uppercase; color: #818cf8;
}
.product-table-wrapper thead th:first-child { width: 90px; text-align: center; }
.product-table-wrapper tbody tr { border-bottom: 1px solid rgba(255,255,255,0.04); transition: background 0.15s; }
.product-table-wrapper tbody tr:hover { background: rgba(255,255,255,0.04); }
.product-table-wrapper tbody tr:last-child { border-bottom: none; }
.product-table-wrapper tbody td {
    padding: 0.65rem 1rem; font-size: 0.84rem; color: #cbd5e1; vertical-align: middle;
}
.product-table-wrapper tbody td:nth-child(2) {
    color: #f1f5f9; max-width: 300px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
}
.product-table-wrapper tbody td:nth-child(3) { font-weight: 700; color: #a5b4fc; white-space: nowrap; }

/* Thumbnail */
.thumb-cell { text-align: center; padding: 0.4rem 0.6rem !important; }
.thumb-img {
    width: 72px; height: 72px; object-fit: contain; border-radius: 10px;
    background: rgba(255,255,255,0.08); padding: 5px;
    border: 1px solid rgba(255,255,255,0.1); display: block; margin: 0 auto;
    transition: transform 0.2s ease;
}
.thumb-img:hover { transform: scale(1.13); cursor: zoom-in; }
.thumb-placeholder {
    width: 72px; height: 72px; border-radius: 10px;
    background: rgba(255,255,255,0.04); border: 1px solid rgba(255,255,255,0.07);
    display: flex; align-items: center; justify-content: center;
    margin: 0 auto; font-size: 1.6rem;
}

/* Badges / links */
.store-badge {
    display: inline-block; font-size: 0.72rem; font-weight: 600; color: #94a3b8;
    background: rgba(255,255,255,0.05); border: 1px solid rgba(255,255,255,0.09);
    padding: 0.17rem 0.52rem; border-radius: 6px; white-space: nowrap;
}
.price-min { color: #34d399 !important; }
.price-max { color: #f472b6 !important; }
.links-cell { display: flex; flex-direction: column; gap: 5px; align-items: flex-start; }
a.offer-link {
    display: inline-flex; align-items: center; gap: 0.25em; font-size: 0.73rem; font-weight: 600;
    color: #818cf8; background: rgba(99,102,241,0.1); border: 1px solid rgba(99,102,241,0.22);
    padding: 0.18rem 0.55rem; border-radius: 6px; text-decoration: none; transition: all 0.15s; white-space: nowrap;
}
a.offer-link:hover { background: rgba(99,102,241,0.22); color: #a5b4fc; }
a.product-link {
    display: inline-flex; align-items: center; gap: 0.25em; font-size: 0.73rem; font-weight: 600;
    color: #6ee7b7; background: rgba(16,185,129,0.08); border: 1px solid rgba(16,185,129,0.2);
    padding: 0.18rem 0.55rem; border-radius: 6px; text-decoration: none; transition: all 0.15s; white-space: nowrap;
}
a.product-link:hover { background: rgba(16,185,129,0.18); color: #a7f3d0; }

/* Pagination */
.pag-info {
    text-align: center; font-size: 0.74rem; color: #4b5563;
    background: rgba(255,255,255,0.03); border: 1px solid rgba(255,255,255,0.06);
    padding: 0.38rem 1rem; border-radius: 8px; margin: 0 auto; width: fit-content;
}

/* Stat cards */
.stat-card {
    background: rgba(255,255,255,0.03); border: 1px solid rgba(255,255,255,0.08);
    border-radius: 14px; padding: 1.25rem 1.5rem; text-align: center;
    position: relative; overflow: hidden; transition: transform 0.2s, border-color 0.2s;
}
.stat-card:hover { transform: translateY(-3px); border-color: rgba(99,102,241,0.3); }
.stat-card::before { content: ''; position: absolute; top: 0; left: 0; right: 0; height: 2px; }
.stat-card.smin::before { background: linear-gradient(90deg, #10b981, #34d399); }
.stat-card.savg::before { background: linear-gradient(90deg, #6366f1, #818cf8); }
.stat-card.smax::before { background: linear-gradient(90deg, #ec4899, #f43f5e); }
.stat-label { font-size: 0.64rem; font-weight: 700; letter-spacing: 0.13em; text-transform: uppercase; color: #64748b; margin-bottom: 0.35rem; }
.stat-value { font-size: 1.45rem; font-weight: 800; letter-spacing: -0.02em; line-height: 1; }
.stat-name { font-size: 0.68rem; color: #475569; margin-top: 0.3rem; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
.stat-card.smin .stat-value { color: #34d399; }
.stat-card.savg .stat-value { color: #818cf8; }
.stat-card.smax .stat-value { color: #f472b6; }

.currency-note {
    font-size: 0.67rem; color: #3f4858; background: rgba(255,255,255,0.03);
    border: 1px solid rgba(255,255,255,0.05); padding: 0.17rem 0.58rem;
    border-radius: 6px; margin-left: 0.4rem; font-weight: 400;
}
.stSpinner > div { border-top-color: #6366f1 !important; }
.js-plotly-plot .plotly, .plot-container { background: transparent !important; }
</style>
""", unsafe_allow_html=True)

# ── Navbar with logo ──────────────────────────────────────────────────────────
col_logo, col_tag = st.columns([1, 4])
with col_logo:
    st.image("buscador_whitelabel/Logo2025_branco 2.png", width=160)
with col_tag:
    st.markdown("""
    <div class="navbar-tag" style="margin-top:14px;display:inline-block;">
        ⚡ Powered by Pesquise Mais · Google Shopping
    </div>
    """, unsafe_allow_html=True)

# ── Hero ──────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="hero-header">
    <h1 class="hero-title">Busca <span>Inteligente</span> de Produtos</h1>
    <p class="hero-sub">Compare preços em tempo real e encontre as melhores ofertas</p>
</div>
""", unsafe_allow_html=True)

# ── Search ────────────────────────────────────────────────────────────────────
st.markdown('<div class="search-wrap">', unsafe_allow_html=True)
col_input, col_btn = st.columns([5, 1])
with col_input:
    q = st.text_input("O que você quer buscar?", value="notebook", key="query_input")
with col_btn:
    # Invisible spacer matching the label height so button aligns perfectly with input
    st.markdown('<span class="btn-label-spacer">.</span>', unsafe_allow_html=True)
    search_clicked = st.button("🔍 Buscar", key="btn_search")
location = "Sao Paulo, State of Sao Paulo, brazil"
st.caption(f"📍 Localização fixa: {location}")
st.markdown('</div>', unsafe_allow_html=True)

# ── Session state ─────────────────────────────────────────────────────────────
ROWS_PER_PAGE = 5
for key, default in [("df_results", None), ("page", 0), ("last_query", ""), ("raw_produtos", [])]:
    if key not in st.session_state:
        st.session_state[key] = default

# ── Excel builder ─────────────────────────────────────────────────────────────
def build_excel(produtos_raw: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    # Styles
    header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill   = PatternFill("solid", fgColor="3730A3")
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align    = Alignment(vertical="center", wrap_text=True)
    thin          = Side(style="thin", color="D1D5DB")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill      = PatternFill("solid", fgColor="F0F0FF")
    price_font    = Font(name="Arial", bold=True, color="4F46E5", size=10)
    link_font     = Font(name="Arial", color="0D6EFD", underline="single", size=9)

    # All available API fields mapped to friendly column names
    columns = [
        ("position",          "Posição"),
        ("title",             "Título"),
        ("price",             "Preço (texto)"),
        ("extracted_price",   "Preço (número)"),
        ("old_price",         "Preço Antigo"),
        ("extracted_old_price","Preço Antigo (número)"),
        ("rating",            "Avaliação"),
        ("reviews",           "Nº Avaliações"),
        ("source",            "Loja"),
        ("store_link",        "Link da Loja"),
        ("link",              "Link Oferta"),
        ("product_link",      "Link Produto"),
        ("product_id",        "ID Produto"),
        ("serpapi_product_api","SerpApi Product API"),
        ("delivery",          "Entrega"),
        ("badge",             "Badge"),
        ("second_hand_condition","Condição (usado)"),
        ("thumbnail",         "URL Thumbnail"),
        ("tag",               "Tag"),
        ("extensions",        "Extensões"),
    ]
    api_keys   = [c[0] for c in columns]
    col_labels = [c[1] for c in columns]

    # Header row
    ws.row_dimensions[1].height = 30
    for col_idx, label in enumerate(col_labels, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        cell.border = border

    # Data rows
    for row_idx, produto in enumerate(produtos_raw, start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        ws.row_dimensions[row_idx].height = 20
        for col_idx, key in enumerate(api_keys, start=1):
            val = produto.get(key, "")
            # extensions list → join as string
            if isinstance(val, list):
                val = " | ".join(str(v) for v in val)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = cell_align
            cell.border    = border
            if fill:
                cell.fill = fill
            # Special formatting
            if key == "extracted_price" and val:
                try:
                    cell.number_format = 'R$ #,##0.00'
                    cell.font = price_font
                except Exception:
                    pass
            if key in ("link", "product_link", "store_link", "serpapi_product_api") and val:
                cell.font = link_font

    # Column widths (auto-ish)
    col_widths = {
        "Posição": 8, "Título": 45, "Preço (texto)": 14, "Preço (número)": 14,
        "Preço Antigo": 14, "Preço Antigo (número)": 18, "Avaliação": 10,
        "Nº Avaliações": 12, "Loja": 20, "Link da Loja": 35, "Link Oferta": 35,
        "Link Produto": 35, "ID Produto": 22, "SerpApi Product API": 35,
        "Entrega": 20, "Badge": 12, "Condição (usado)": 18,
        "URL Thumbnail": 45, "Tag": 12, "Extensões": 30,
    }
    for col_idx, label in enumerate(col_labels, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(label, 18)

    # Freeze header
    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Resumo")
    prices = [p.get("extracted_price") for p in produtos_raw if p.get("extracted_price") is not None]
    if prices:
        min_p   = min(prices)
        max_p   = max(prices)
        avg_p   = sum(prices) / len(prices)
        min_prod = next((p.get("title","") for p in produtos_raw if p.get("extracted_price") == min_p), "")
        max_prod = next((p.get("title","") for p in produtos_raw if p.get("extracted_price") == max_p), "")
        summary_rows = [
            ("Métrica", "Valor", "Produto"),
            ("Total de resultados", len(produtos_raw), ""),
            ("Menor preço", min_p, min_prod),
            ("Maior preço", max_p, max_prod),
            ("Preço médio", avg_p, ""),
        ]
        for r_idx, row in enumerate(summary_rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                    cell.fill = PatternFill("solid", fgColor="3730A3")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx == 2 and r_idx > 2:
                    cell.number_format = 'R$ #,##0.00'
                    cell.font = Font(name="Arial", bold=True, color="4F46E5")
                cell.border = border
        for col_l, width in [("A", 22), ("B", 18), ("C", 50)]:
            ws2.column_dimensions[col_l].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ── API ───────────────────────────────────────────────────────────────────────
api_key = os.environ.get("SERP_APIKEY")
if not api_key:
    st.warning("A variável de ambiente SERP_APIKEY não está definida.")
else:
    if search_clicked:
        with st.spinner("Buscando resultados..."):
            params = {
                "engine": "google_shopping",
                "q": q,
                "api_key": api_key,
                "location": "Sao Paulo, State of Sao Paulo, brazil",
                "google_domain": "google.com.br",
                "gl": "br",
                "hl": "pt",
                "device": "desktop",
                "sort_by": 1,
                "no_cache": True
            }
            if location.strip():
                params["location"] = location.strip()
            resp = requests.get("https://serpapi.com/search.json", params=params)

            if resp.status_code == 200:
                data = resp.json()
                produtos = data.get("shopping_results", [])
                if not produtos:
                    st.warning("Nenhum resultado encontrado.")
                    st.session_state.df_results  = None
                    st.session_state.raw_produtos = []
                else:
                    def parse_price(price):
                        if isinstance(price, (int, float)):
                            return float(price)
                        if isinstance(price, str):
                            price = price.replace("R$", "").replace("$", "").replace(",", ".").strip()
                            try:    return float(price)
                            except: return None
                        return None

                    df = pd.DataFrame([
                        {
                            "Título":       p.get("title"),
                            "Preço":        float(p["extracted_price"]) if p.get("extracted_price") is not None else parse_price(p.get("price", 0)),
                            "Loja":         p.get("source"),
                            "Link":         p.get("link"),
                            "ProductLink":  p.get("product_link", ""),
                            "Thumbnail":    p.get("thumbnail", ""),
                        }
                        for p in produtos if p.get("price")
                    ])
                    st.session_state.df_results   = df
                    st.session_state.raw_produtos  = produtos
                    st.session_state.page          = 0
                    st.session_state.last_query    = q
            else:
                st.error(f"Erro na consulta: {resp.status_code}")

    # ── Render ────────────────────────────────────────────────────────────────
    if st.session_state.df_results is not None:
        df    = st.session_state.df_results
        total = len(df)
        total_pages = max(1, -(-total // ROWS_PER_PAGE))
        page  = st.session_state.page
        start = page * ROWS_PER_PAGE
        end   = start + ROWS_PER_PAGE
        df_page = df.iloc[start:end].copy()

        global_min = df["Preço"].min()
        global_max = df["Preço"].max()
        min_loja   = df.loc[df["Preço"].idxmin(), "Loja"]
        max_loja   = df.loc[df["Preço"].idxmax(), "Loja"]

        # ── Stat cards ───────────────────────────────────────────────────
        st.markdown('<div class="section-heading">📈 Resumo de preços</div>', unsafe_allow_html=True)
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(f"""
            <div class="stat-card smin">
                <div class="stat-label">▼ Menor preço</div>
                <div class="stat-value">R$ {global_min:,.2f}</div>
                <div class="stat-name">🏪 {min_loja or '—'}</div>
            </div>""", unsafe_allow_html=True)
        with col2:
            st.markdown(f"""
            <div class="stat-card savg">
                <div class="stat-label">◆ Média geral</div>
                <div class="stat-value">R$ {df['Preço'].mean():,.2f}</div>
                <div class="stat-name">{total} produtos encontrados</div>
            </div>""", unsafe_allow_html=True)
        with col3:
            st.markdown(f"""
            <div class="stat-card smax">
                <div class="stat-label">▲ Maior preço</div>
                <div class="stat-value">R$ {global_max:,.2f}</div>
                <div class="stat-name">🏪 {max_loja or '—'}</div>
            </div>""", unsafe_allow_html=True)

        # ── Export Excel button ───────────────────────────────────────────
        st.markdown('<div class="section-heading">📥 Exportar dados</div>', unsafe_allow_html=True)
        excel_bytes = build_excel(st.session_state.raw_produtos)
        fname = f"pesquisemais_{st.session_state.last_query.replace(' ','_')}.xlsx"
        st.download_button(
            label="⬇️ Baixar Excel completo",
            data=excel_bytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help=f"Exporta todos os {total} resultados com todos os campos da API"
        )

        # ── Table ─────────────────────────────────────────────────────────
        st.markdown(f"""
        <div class="section-heading">
            🛒 Resultados
            <span class="currency-note">💡 Preços em R$ — busca localizada BR</span>
        </div>
        <div class="results-badge">✦ {total} produtos para "{st.session_state.last_query}" · página {page+1}/{total_pages}</div>
        """, unsafe_allow_html=True)

        def make_thumb(url):
            if url:
                return (f'<div class="thumb-cell"><img src="{url}" class="thumb-img" loading="lazy" '
                        f'onerror="this.parentElement.innerHTML=\'<div class=\\\"thumb-placeholder\\\">📦</div>\'"></div>')
            return '<div class="thumb-cell"><div class="thumb-placeholder">📦</div></div>'

        def make_price_cell(row):
            val = row["Preço"]
            if pd.isnull(val): return "–"
            fmt = f"R$ {val:,.2f}"
            if val == global_min: return f'<span class="price-min">▼ {fmt}</span>'
            if val == global_max: return f'<span class="price-max">▲ {fmt}</span>'
            return fmt

        def make_links(row):
            parts = []
            if row.get("Link"):
                parts.append(f'<a href="{row["Link"]}" target="_blank" class="offer-link">🔗 Ver oferta</a>')
            if row.get("ProductLink"):
                parts.append(f'<a href="{row["ProductLink"]}" target="_blank" class="product-link">🛒 Produto</a>')
            return f'<div class="links-cell">{"".join(parts)}</div>' if parts else "–"

        df_page["Foto"]      = df_page["Thumbnail"].apply(make_thumb)
        df_page["Preço_fmt"] = df_page.apply(make_price_cell, axis=1)
        df_page["Loja_fmt"]  = df_page["Loja"].apply(lambda l: f'<span class="store-badge">{l}</span>' if l else "–")
        df_page["Links"]     = df_page.apply(make_links, axis=1)

        table_html = df_page.to_html(
            columns=["Foto", "Título", "Preço_fmt", "Loja_fmt", "Links"],
            escape=False, index=False,
            header=["", "Produto", "Preço", "Loja", "Links"]
        )
        st.write(f'<div class="product-table-wrapper">{table_html}</div>', unsafe_allow_html=True)

        # ── Pagination ────────────────────────────────────────────────────
        cp, ci, cn = st.columns([1, 3, 1])
        with cp:
            if st.button("← Anterior", disabled=(page == 0), key="btn_prev"):
                st.session_state.page -= 1
                st.rerun()
        with ci:
            st.markdown(
                f'<div class="pag-info">Página {page+1} de {total_pages} &nbsp;·&nbsp; itens {start+1}–{min(end,total)} de {total}</div>',
                unsafe_allow_html=True)
        with cn:
            if st.button("Próxima →", disabled=(page >= total_pages - 1), key="btn_next"):
                st.session_state.page += 1
                st.rerun()

        # ── Chart ─────────────────────────────────────────────────────────
        st.markdown('<div class="section-heading" style="margin-top:2rem">📊 Preços por loja</div>', unsafe_allow_html=True)
        df_chart = df.copy()
        df_chart["Preço_fmt"] = df_chart["Preço"].apply(lambda x: f"R$ {x:,.2f}" if pd.notnull(x) else "–")
        fig = px.bar(df_chart, x="Loja", y="Preço", color="Loja", text="Preço_fmt",
                     labels={"Preço": "Preço (R$)", "Loja": "Loja"})
        fig.update_traces(textposition="outside")
        fig.update_layout(
            showlegend=False, height=420,
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            font=dict(family="Open Sans, sans-serif", color="#94a3b8", size=12),
            xaxis=dict(gridcolor="rgba(255,255,255,0.05)", linecolor="rgba(255,255,255,0.08)"),
            yaxis=dict(gridcolor="rgba(255,255,255,0.05)", linecolor="rgba(255,255,255,0.08)"),
            colorway=["#6366f1","#818cf8","#a78bfa","#ec4899","#f43f5e","#10b981","#0ea5e9","#f59e0b"],
            bargap=0.35, margin=dict(t=20),
        )
        st.plotly_chart(fig, use_container_width=True)